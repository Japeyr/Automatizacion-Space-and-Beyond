from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import time
from threading import Lock

lock = Lock()


def head_planilla():
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Reporte de Defectos"

    encabezados = [
        "ID", "Título", "Precondiciones", "Datos de Entrada",
        "Pasos", "Resultado Esperado", "Estado",
        "Resultado Obtenido", "Tester", "Fecha"
    ]
    hoja.append(encabezados)

    # Estilos
    alineacion = Alignment(horizontal="center", vertical="center")
    fuente_negrita = Font(bold=True)
    fondo = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Celeste claro

    # Aplicar estilos a cada celda del encabezado
    for col_index in range(1, len(encabezados) + 1):
        celda = hoja.cell(row=1, column=col_index)
        celda.alignment = alineacion
        celda.font = fuente_negrita
        celda.fill = fondo

    # Definir estilo de borde
    borde_fino = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar a todas las celdas con datos (desde fila 1 hasta la última)
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, max_col=hoja.max_column):
        for celda in fila:
            celda.border = borde_fino

    wb.save("Casos_Prueba_Space&Beyond.xlsx")


def datos1(fila, user, password, resultado):
    caso_prueba = f"CP{fila}"
    titulo = "Loguearse en el Sitio"
    precondiciones = "Ingresar al sitio\n https://demo.testim.io/"

    usuario = user if user else "vacio"
    contrasenia = password if password else "vacio"

    pasos = '''1 - Click en botón login\n
2 - Introducir Usuario\n
3 - Introducir Contraseña\n
4 - Click en botón Submit\n'''

    if resultado == "esperado":
        resultado_esperado = "✅ Resultado esperado"
        resultado_obtenido = "OK"
        estado = "Aprobado"
    else:
        resultado_esperado = "❌ Resultado NO esperado"
        resultado_obtenido = "No OK"
        estado = "Falla"

    tester = 'Jorge Peyrano'
    # Obtener fecha y hora actual
    fecha_hora_actual = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    print("📅 Fecha y hora actual:", fecha_hora_actual)

    wb = load_workbook("Casos_Prueba_Space&Beyond.xlsx")

    hoja = wb.active

    hoja.append([
        caso_prueba,
        titulo,
        precondiciones,
        f"Usuario: {usuario}\n Contraseña: {contrasenia}",
        pasos,
        resultado_esperado,
        estado,
        resultado_obtenido,
        tester,
        fecha_hora_actual
    ])

    # Columnas a centrar (excepto la 5 = "Pasos")
    columnas_a_centrar = [1, 2, 6, 7, 8, 9, 10]

    for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row):
        for col_index in columnas_a_centrar:
            celda = fila[col_index - 1]  # -1 porque los índices empiezan en 0
            celda.alignment = Alignment(horizontal="center", vertical="center")

    # Celda de "Datos de Precondiciones" (columna 3)
    celda_precondiones = hoja.cell(row=hoja.max_row, column=3)
    celda_precondiones.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Celda de "Datos de Entrada" (columna 4)
    celda_datos = hoja.cell(row=hoja.max_row, column=4)
    celda_datos.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Celda de "Pasos" (columna 5)
    celda_pasos = hoja.cell(row=hoja.max_row, column=5)
    celda_pasos.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Supongamos que la columna de "Pasos" es la número 5
    col_id = get_column_letter(1)
    col_titulo = get_column_letter(2)
    col_precondiciones = get_column_letter(3)
    col_datos_entrada = get_column_letter(4)
    col_pasos = get_column_letter(5)  # Devuelve 'E'
    col_resultado_esperado = get_column_letter(6)
    col_estado = get_column_letter(7)
    col_resultado_obtenido = get_column_letter(8)
    col_tester = get_column_letter(9)
    col_fecha = get_column_letter(10)

    # Ajustar el ancho de la columna
    hoja.column_dimensions[col_id].width =10
    hoja.column_dimensions[col_titulo].width = 25
    hoja.column_dimensions[col_precondiciones].width = 25
    hoja.column_dimensions[col_datos_entrada].width = 25
    hoja.column_dimensions[col_pasos].width = 30  # Ancho en caracteres
    hoja.column_dimensions[col_resultado_esperado].width = 25
    hoja.column_dimensions[col_estado].width = 10
    hoja.column_dimensions[col_resultado_obtenido].width = 20
    hoja.column_dimensions[col_tester].width = 15
    hoja.column_dimensions[col_fecha].width = 20

    # Definir estilo de borde
    borde_fino = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar a todas las celdas con datos (desde fila 1 hasta la última)
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, max_col=hoja.max_column):
        for celda in fila:
            celda.border = borde_fino

    wb.save("Casos_Prueba_Space&Beyond.xlsx")

def datos2(fila_logout, resultado):
    caso_prueba = f"CP{fila_logout-1}"
    titulo = "Desloguearse en el Sitio"
    precondiciones = '''1- Haber ingresado al sitio\n https://demo.testim.io/\n 
    2 - Haber hecho Click en botón login\n
    3 - Haber Introducido el Usuario\n
    4 - Haber Introducido la Contraseña\n
    5 - Haber hecho Click en botón Submit\n'''

    datos_entrada = "No aplica"

    pasos = '''1 - Click en botón logout\n'''

    if resultado == "Esperado":
        resultado_esperado = "✅ Resultado Esperado"
        resultado_obtenido = "OK"
        estado = "Aprobado"
    else:
        resultado_esperado = "❌ Resultado NO Esperado"
        resultado_obtenido = "No OK"
        estado = "Falla"

    tester = 'Jorge Peyrano'
    # Obtener fecha y hora actual
    fecha_hora_actual = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    print("📅 Fecha y hora actual:", fecha_hora_actual)

    wb = load_workbook("Casos_Prueba_Space&Beyond.xlsx")

    hoja = wb.active

    hoja.append([
        caso_prueba,
        titulo,
        precondiciones,
        datos_entrada,
        pasos,
        resultado_esperado,
        estado,
        resultado_obtenido,
        tester,
        fecha_hora_actual
    ])

    # Columnas a centrar (excepto la 5 = "Pasos")
    columnas_a_centrar = [1, 2, 6, 7, 8, 9, 10]

    for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row):
        for col_index in columnas_a_centrar:
            celda = fila[col_index - 1]  # -1 porque los índices empiezan en 0
            celda.alignment = Alignment(horizontal="center", vertical="center")

    # Celda de "Datos de Precondiciones" (columna 3)
    celda_precondiones = hoja.cell(row=hoja.max_row, column=3)
    celda_precondiones.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Celda de "Datos de Entrada" (columna 4)
    celda_datos_entrada = hoja.cell(row=hoja.max_row, column=4)
    celda_datos_entrada.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Celda de "Pasos" (columna 5)
    celda_pasos = hoja.cell(row=hoja.max_row, column=5)
    celda_pasos.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Supongamos que la columna de "Pasos" es la número 5
    col_id = get_column_letter(1)
    col_titulo = get_column_letter(2)
    col_precondiciones = get_column_letter(3)
    col_datos_entrada = get_column_letter(4)
    col_pasos = get_column_letter(5)  # Devuelve 'E'
    col_resultado_esperado = get_column_letter(6)
    col_estado = get_column_letter(7)
    col_resultado_obtenido = get_column_letter(8)
    col_tester = get_column_letter(9)
    col_fecha = get_column_letter(10)

    # Ajustar el ancho de la columna
    hoja.column_dimensions[col_id].width = 10
    hoja.column_dimensions[col_titulo].width = 25
    hoja.column_dimensions[col_precondiciones].width = 25
    hoja.column_dimensions[col_datos_entrada].width = 25
    hoja.column_dimensions[col_pasos].width = 30  # Ancho en caracteres
    hoja.column_dimensions[col_resultado_esperado].width = 25
    hoja.column_dimensions[col_estado].width = 10
    hoja.column_dimensions[col_resultado_obtenido].width = 20
    hoja.column_dimensions[col_tester].width = 15
    hoja.column_dimensions[col_fecha].width = 20

    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    celda_pasos_logout = hoja.cell(row=hoja.max_row, column=5)
    celda_pasos_logout.alignment = alineacion_centro

    # Definir estilo de borde
    borde_fino = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar a todas las celdas con datos (desde fila 1 hasta la última)
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, max_col=hoja.max_column):
        for celda in fila:
            celda.border = borde_fino

    wb.save("Casos_Prueba_Space&Beyond.xlsx")

# Contenido a agregar al final de Planilla_Calculo.py
def datos3(fila_id, planeta_num, nombre, email, ss, tel, codigo, resultado_validacion, resultado_web):
    wb = load_workbook("Casos_Prueba_Space&Beyond.xlsx")

    hoja = wb.active
    secuencia_id = hoja.max_row

    caso_prueba = f"CP{secuencia_id}"
    titulo = f"Reserva Planeta {planeta_num}"
    precondiciones = '''1- Logueado en el sitio\n https://demo.testim.io/\n
    2 - En la página de planetas'''

    datos_entrada = f"Nombre: {nombre}\nEmail: {email}\nSS: {ss}\nTeléfono: {tel}\nCódigo: {codigo}"

    pasos = '''1 - Click en botón 'Book Now' del planeta\n
2 - Completar formulario de reserva\n
3 - Click en 'Apply' para el código\n
4 - Aceptar términos y condiciones\n
5 - Verificar estado del botón 'Pay now' '''

    if resultado_validacion == "OK" and resultado_web == "OK":
        resultado_esperado = "✅ Datos Válidos (Web OK)"
        resultado_obtenido = "APROBADO"
        estado = "Aprobado"
    elif resultado_validacion == "FALLÓ_LOGICA":
        resultado_esperado = "❌ Datos Inválidos detectados por lógica"
        resultado_obtenido = "FALLÓ"
        estado = "Falla (Datos)"
    elif resultado_validacion == "OK" and resultado_web == "FALLÓ_WEB":
        resultado_esperado = "✅ Datos Válidos (Web FALLÓ)"
        resultado_obtenido = "FALLÓ"
        estado = "Falla (Web)"
    else:
        resultado_esperado = "⚠️ Error durante la ejecución"
        resultado_obtenido = "ERROR"
        estado = "Error"

    tester = 'Jorge Peyrano'
    fecha_hora_actual = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    print("📅 Fecha y hora actual:", fecha_hora_actual)

    wb = load_workbook("Casos_Prueba_Space&Beyond.xlsx")
    hoja = wb.active

    hoja.append([
        caso_prueba,
        titulo,
        precondiciones,
        datos_entrada,
        pasos,
        resultado_esperado,
        estado,
        resultado_obtenido,
        tester,
        fecha_hora_actual
    ])

    # Se aplica el mismo formateo que en datos1 y datos2 (ajustes de celda y borde)

    # Columnas a centrar
    columnas_a_centrar = [1, 2, 6, 7, 8, 9, 10]

    for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row):
        for col_index in columnas_a_centrar:
            celda = fila[col_index - 1]
            celda.alignment = Alignment(horizontal="center", vertical="center")

    # Wrap text y alineación para columnas específicas
    for col_index in [3, 4]:
        celda = hoja.cell(row=hoja.max_row, column=col_index)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    celda_pasos = hoja.cell(row=hoja.max_row, column=5)
    celda_pasos.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Ajustar ancho de columnas (se usa el mismo que ya tienes)
    col_id = get_column_letter(1)
    col_titulo = get_column_letter(2)
    col_precondiciones = get_column_letter(3)
    col_datos_entrada = get_column_letter(4)
    col_pasos = get_column_letter(5)
    col_resultado_esperado = get_column_letter(6)
    col_estado = get_column_letter(7)
    col_resultado_obtenido = get_column_letter(8)
    col_tester = get_column_letter(9)
    col_fecha = get_column_letter(10)

    hoja.column_dimensions[col_id].width = 10
    hoja.column_dimensions[col_titulo].width = 25
    hoja.column_dimensions[col_precondiciones].width = 25
    hoja.column_dimensions[col_datos_entrada].width = 25
    hoja.column_dimensions[col_pasos].width = 30
    hoja.column_dimensions[col_resultado_esperado].width = 25
    hoja.column_dimensions[col_estado].width = 10
    hoja.column_dimensions[col_resultado_obtenido].width = 20
    hoja.column_dimensions[col_tester].width = 15
    hoja.column_dimensions[col_fecha].width = 20

    # Definir estilo de borde
    borde_fino = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar a todas las celdas con datos
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, max_col=hoja.max_column):
        for celda in fila:
            celda.border = borde_fino

    wb.save("Casos_Prueba_Space&Beyond.xlsx")

